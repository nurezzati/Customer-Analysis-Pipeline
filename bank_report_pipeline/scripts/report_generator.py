"""
report_generator.py
Builds a professional Excel report from analysis results.
Sheets: Cover, KPI Dashboard, Monthly Trend, Branch Analysis,
        Transaction Breakdown, Segment Analysis, Top Customers, Anomaly Report
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
from datetime import datetime


# ── STYLE CONSTANTS ───────────────────────────────────────────────────────────

CUSTOMER_YELLOW = "FFCC00"
CUSTOMER_DARK   = "1A1A1A"
HEADER_BLUE    = "1F3864"
ACCENT_TEAL    = "2E75B6"
LIGHT_BLUE     = "D6E4F0"
LIGHT_YELLOW   = "FFF9D6"
LIGHT_GRAY     = "F5F5F5"
WHITE          = "FFFFFF"
RED_ALERT      = "C00000"
GREEN_OK       = "375623"

def thin_border():
    s = Side(border_style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def header_fill(color=HEADER_BLUE):
    return PatternFill("solid", fgColor=color)

def cell_fill(color):
    return PatternFill("solid", fgColor=color)

def hdr_font(color=WHITE, size=11, bold=True):
    return Font(name="Arial", size=size, bold=bold, color=color)

def body_font(size=10, bold=False, color=CUSTOMER_DARK):
    return Font(name="Arial", size=size, bold=bold, color=color)

def currency_fmt():
    return '#,##0.00'

def set_col_widths(ws, widths: dict):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# ── HELPERS ───────────────────────────────────────────────────────────────────

def write_table_header(ws, row, cols, fill_color=HEADER_BLUE, font_color=WHITE):
    for c, label in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=label)
        cell.font      = hdr_font(color=font_color)
        cell.fill      = header_fill(fill_color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = thin_border()

def write_data_rows(ws, start_row, data, fmt_map=None, number_cols=None):
    for r_idx, row_data in enumerate(data):
        row_num = start_row + r_idx
        bg = LIGHT_GRAY if r_idx % 2 == 0 else WHITE
        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=c_idx, value=val)
            cell.fill   = cell_fill(bg)
            cell.font   = body_font()
            cell.border = thin_border()
            cell.alignment = Alignment(horizontal="right" if isinstance(val, (int, float)) else "left",
                                       vertical="center")
            if fmt_map and c_idx in fmt_map:
                cell.number_format = fmt_map[c_idx]

def kpi_card(ws, row, col, label, value, note="", fill=LIGHT_BLUE):
    lbl = ws.cell(row=row, column=col, value=label)
    lbl.font      = Font(name="Arial", size=9, bold=True, color="555555")
    lbl.fill      = cell_fill(fill)
    lbl.alignment = Alignment(horizontal="center", vertical="center")

    val = ws.cell(row=row+1, column=col, value=value)
    val.font      = Font(name="Arial", size=16, bold=True, color=HEADER_BLUE)
    val.fill      = cell_fill(fill)
    val.alignment = Alignment(horizontal="center", vertical="center")

    if note:
        n = ws.cell(row=row+2, column=col, value=note)
        n.font      = Font(name="Arial", size=8, color="888888")
        n.fill      = cell_fill(fill)
        n.alignment = Alignment(horizontal="center")

    for r in range(row, row+3):
        ws.cell(r, col).border = thin_border()


# ── SHEET BUILDERS ────────────────────────────────────────────────────────────

def build_cover(wb, kpi):
    ws = wb.active
    ws.title = "Cover"
    ws.sheet_view.showGridLines = False

    # Banner
    ws.merge_cells("A1:J3")
    banner = ws["A1"]
    banner.value     = "CUSTOMER  |  TRANSACTION ANALYTICS REPORT"
    banner.font      = Font(name="Arial", size=20, bold=True, color=CUSTOMER_DARK)
    banner.fill      = cell_fill(CUSTOMER_YELLOW)
    banner.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 30
    ws.row_dimensions[3].height = 20

    # Sub-header
    ws.merge_cells("A4:J4")
    sub = ws["A4"]
    sub.value     = f"Full Year 2024  ·  Generated: {kpi['report_date']}"
    sub.font      = Font(name="Arial", size=12, color="444444")
    sub.alignment = Alignment(horizontal="center", vertical="center")
    sub.fill      = cell_fill("FFFDE7")
    ws.row_dimensions[4].height = 22

    ws.merge_cells("A6:J6")
    ws["A6"].value     = "EXECUTIVE SUMMARY"
    ws["A6"].font      = hdr_font(color=HEADER_BLUE, size=13)
    ws["A6"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[6].height = 22

    # KPI cards across row 8-10
    kpi_data = [
        ("Total Transactions",  f"{kpi['total_transactions']:,}",    "Completed & Pending"),
        ("Total Volume (RM)",   f"{kpi['total_volume']:,.2f}",        "Full year 2024"),
        ("Avg Transaction (RM)",f"{kpi['avg_transaction']:,.2f}",     "Per transaction"),
        ("Unique Customers",    f"{kpi['unique_customers']:,}",       "Active customers"),
        ("Anomalies Flagged",   str(kpi['anomalies_flagged']),        "High-value outliers"),
    ]
    fills = [LIGHT_BLUE, LIGHT_YELLOW, LIGHT_BLUE, LIGHT_YELLOW, "FFE0E0"]
    cols  = [2, 4, 6, 8, 10]
    for (label, val, note), fill, col in zip(kpi_data, fills, cols):
        ws.row_dimensions[8].height  = 18
        ws.row_dimensions[9].height  = 32
        ws.row_dimensions[10].height = 16
        kpi_card(ws, 8, col, label, val, note, fill)

    # Navigation guide
    ws.merge_cells("A12:J12")
    ws["A12"].value     = "REPORT CONTENTS"
    ws["A12"].font      = hdr_font(color=HEADER_BLUE, size=11)
    ws["A12"].alignment = Alignment(horizontal="left")
    ws.row_dimensions[12].height = 20

    contents = [
        ("Monthly Trend",         "Monthly transaction volume and count — Jan to Dec 2024"),
        ("Branch Analysis",       "Volume and customer breakdown by branch"),
        ("Transaction Breakdown", "Split by transaction type and product"),
        ("Segment Analysis",      "Retail, SME, and Wealth Management performance"),
        ("Top Customers",         "Top 10 customers by total transaction volume"),
        ("Anomaly Report",        "Transactions flagged as statistical outliers (>3σ)"),
    ]
    for i, (sheet, desc) in enumerate(contents, 13):
        ws.cell(i, 2, value=sheet).font  = Font(name="Arial", size=10, bold=True, color=ACCENT_TEAL)
        ws.cell(i, 3, value=desc).font   = body_font()
        ws.row_dimensions[i].height = 16

    set_col_widths(ws, {"A": 2, "B": 5, "C": 2, "D": 2, "E": 2,
                        "F": 2, "G": 2, "H": 2, "I": 2, "J": 2})
    for col in "BCDEFGHIJ":
        ws.column_dimensions[col].width = 18


def build_monthly(wb, monthly_df):
    ws = wb.create_sheet("Monthly Trend")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:E1")
    ws["A1"].value     = "Monthly Transaction Trend — 2024"
    ws["A1"].font      = hdr_font(color=HEADER_BLUE, size=13)
    ws["A1"].fill      = cell_fill(LIGHT_YELLOW)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    cols = ["Month", "Transactions", "Volume (RM)", "MoM Volume Change (%)", "MoM Txn Change (%)"]
    write_table_header(ws, 3, cols)
    ws.row_dimensions[3].height = 28

    rows = []
    for i, r in monthly_df.iterrows():
        rows.append([r["month_str"], r["transactions"], r["volume"], None, None])

    write_data_rows(ws, 4, rows, fmt_map={3: currency_fmt()})

    # MoM formulas
    for r in range(5, 4 + len(rows)):
        ws.cell(r, 4).value        = f"=IF(C{r-1}=0,\"-\",(C{r}-C{r-1})/C{r-1})"
        ws.cell(r, 4).number_format = "0.0%"
        ws.cell(r, 4).font         = body_font()
        ws.cell(r, 4).fill         = cell_fill(LIGHT_GRAY if (r - 4) % 2 == 0 else WHITE)
        ws.cell(r, 4).border       = thin_border()
        ws.cell(r, 4).alignment    = Alignment(horizontal="right")

        ws.cell(r, 5).value        = f"=IF(B{r-1}=0,\"-\",(B{r}-B{r-1})/B{r-1})"
        ws.cell(r, 5).number_format = "0.0%"
        ws.cell(r, 5).font         = body_font()
        ws.cell(r, 5).fill         = cell_fill(LIGHT_GRAY if (r - 4) % 2 == 0 else WHITE)
        ws.cell(r, 5).border       = thin_border()
        ws.cell(r, 5).alignment    = Alignment(horizontal="right")

    # Totals row
    tr = 4 + len(rows)
    ws.cell(tr, 1, "TOTAL").font  = hdr_font(color=WHITE)
    ws.cell(tr, 1).fill            = header_fill(HEADER_BLUE)
    ws.cell(tr, 1).border          = thin_border()
    ws.cell(tr, 2).value           = f"=SUM(B4:B{tr-1})"
    ws.cell(tr, 2).font            = hdr_font(color=WHITE)
    ws.cell(tr, 2).fill            = header_fill(HEADER_BLUE)
    ws.cell(tr, 2).border          = thin_border()
    ws.cell(tr, 2).alignment       = Alignment(horizontal="right")
    ws.cell(tr, 3).value           = f"=SUM(C4:C{tr-1})"
    ws.cell(tr, 3).number_format   = currency_fmt()
    ws.cell(tr, 3).font            = hdr_font(color=WHITE)
    ws.cell(tr, 3).fill            = header_fill(HEADER_BLUE)
    ws.cell(tr, 3).border          = thin_border()
    ws.cell(tr, 3).alignment       = Alignment(horizontal="right")

    # Bar chart — volume
    chart = BarChart()
    chart.type    = "col"
    chart.title   = "Monthly Transaction Volume (RM)"
    chart.y_axis.title = "Volume (RM)"
    chart.x_axis.title = "Month"
    chart.style   = 10
    chart.width   = 22
    chart.height  = 12
    data  = Reference(ws, min_col=3, min_row=3, max_row=3+len(rows))
    cats  = Reference(ws, min_col=1, min_row=4, max_row=3+len(rows))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.series[0].graphicalProperties.solidFill = ACCENT_TEAL
    ws.add_chart(chart, f"G3")

    set_col_widths(ws, {"A": 14, "B": 15, "C": 18, "D": 22, "E": 20})


def build_branch(wb, branch_df):
    ws = wb.create_sheet("Branch Analysis")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:F1")
    ws["A1"].value     = "Branch Performance — Full Year 2024"
    ws["A1"].font      = hdr_font(color=HEADER_BLUE, size=13)
    ws["A1"].fill      = cell_fill(LIGHT_YELLOW)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    cols = ["Branch", "Transactions", "Total Volume (RM)", "Unique Customers", "Avg Txn (RM)", "Volume Share (%)"]
    write_table_header(ws, 3, cols)

    rows = []
    total_vol = branch_df["volume"].sum()
    for _, r in branch_df.iterrows():
        rows.append([
            r["branch"], r["transactions"], r["volume"],
            r["customers"], r["avg_txn"],
            round(r["volume"] / total_vol * 100, 1)
        ])
    write_data_rows(ws, 4, rows, fmt_map={3: currency_fmt(), 5: currency_fmt(), 6: "0.0\"%\""})

    # Chart
    chart = BarChart()
    chart.type    = "bar"
    chart.title   = "Total Volume by Branch"
    chart.y_axis.title = "Branch"
    chart.x_axis.title = "Volume (RM)"
    chart.style   = 10
    chart.width   = 20
    chart.height  = 12
    data = Reference(ws, min_col=3, min_row=3, max_row=3+len(rows))
    cats = Reference(ws, min_col=1, min_row=4, max_row=3+len(rows))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.series[0].graphicalProperties.solidFill = CUSTOMER_YELLOW
    ws.add_chart(chart, "H3")

    set_col_widths(ws, {"A": 20, "B": 16, "C": 22, "D": 18, "E": 18, "F": 18})


def build_txn_breakdown(wb, by_type_df, by_product_df):
    ws = wb.create_sheet("Transaction Breakdown")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:E1")
    ws["A1"].value     = "Transaction Breakdown — By Type & Product"
    ws["A1"].font      = hdr_font(color=HEADER_BLUE, size=13)
    ws["A1"].fill      = cell_fill(LIGHT_YELLOW)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    # By type
    ws.cell(3, 1, "BY TRANSACTION TYPE").font = hdr_font(color=HEADER_BLUE, size=11)
    write_table_header(ws, 4, ["Transaction Type", "Count", "Volume (RM)", "Share (%)"])
    rows = [[r["transaction_type"], r["transactions"], r["volume"], r["share_pct"]]
            for _, r in by_type_df.iterrows()]
    write_data_rows(ws, 5, rows, fmt_map={3: currency_fmt()})

    # By product (offset right)
    ws.cell(3, 6, "BY PRODUCT").font = hdr_font(color=HEADER_BLUE, size=11)
    write_table_header(ws, 4, ["Product", "Count", "Volume (RM)"], fill_color=ACCENT_TEAL)
    for i, (_, r) in enumerate(by_product_df.iterrows()):
        row_num = 5 + i
        bg = LIGHT_GRAY if i % 2 == 0 else WHITE
        for c, val in enumerate([r["product"], r["transactions"], r["volume"]], 6):
            cell = ws.cell(row_num, c, val)
            cell.fill   = cell_fill(bg)
            cell.font   = body_font()
            cell.border = thin_border()
            cell.alignment = Alignment(horizontal="right" if isinstance(val, (int, float)) else "left")
            if c == 8:
                cell.number_format = currency_fmt()

    # Pie chart — by type
    pie = PieChart()
    pie.title  = "Transaction Mix by Type"
    pie.style  = 10
    pie.width  = 16
    pie.height = 12
    data = Reference(ws, min_col=3, min_row=4, max_row=4+len(rows))
    cats = Reference(ws, min_col=1, min_row=5, max_row=4+len(rows))
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(cats)
    ws.add_chart(pie, "A14")

    set_col_widths(ws, {"A": 22, "B": 12, "C": 20, "D": 12,
                        "E": 4,  "F": 24, "G": 12, "H": 20})


def build_segment(wb, seg_df):
    ws = wb.create_sheet("Segment Analysis")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:E1")
    ws["A1"].value     = "Customer Segment Analysis — 2024"
    ws["A1"].font      = hdr_font(color=HEADER_BLUE, size=13)
    ws["A1"].fill      = cell_fill(LIGHT_YELLOW)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    cols = ["Segment", "Transactions", "Total Volume (RM)", "Unique Customers", "Avg Volume/Customer (RM)"]
    write_table_header(ws, 3, cols)
    rows = [[r["segment"], r["transactions"], r["volume"], r["customers"], r["avg_per_cust"]]
            for _, r in seg_df.iterrows()]
    write_data_rows(ws, 4, rows, fmt_map={3: currency_fmt(), 5: currency_fmt()})

    # Grouped bar chart
    chart = BarChart()
    chart.type    = "col"
    chart.title   = "Volume vs Transactions by Segment"
    chart.grouping = "clustered"
    chart.style   = 10
    chart.width   = 22
    chart.height  = 12
    vol  = Reference(ws, min_col=3, min_row=3, max_row=3+len(rows))
    cats = Reference(ws, min_col=1, min_row=4, max_row=3+len(rows))
    chart.add_data(vol, titles_from_data=True)
    chart.set_categories(cats)
    chart.series[0].graphicalProperties.solidFill = ACCENT_TEAL
    ws.add_chart(chart, "G3")

    set_col_widths(ws, {"A": 22, "B": 16, "C": 22, "D": 20, "E": 26})


def build_top_customers(wb, top_df):
    ws = wb.create_sheet("Top Customers")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:D1")
    ws["A1"].value     = "Top 10 Customers by Transaction Volume — 2024"
    ws["A1"].font      = hdr_font(color=HEADER_BLUE, size=13)
    ws["A1"].fill      = cell_fill(LIGHT_YELLOW)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    cols = ["Rank", "Customer ID", "Transactions", "Total Volume (RM)"]
    write_table_header(ws, 3, cols)

    rows = []
    for i, (_, r) in enumerate(top_df.iterrows(), 1):
        rows.append([i, r["customer_id"], r["transactions"], r["total_volume"]])
    write_data_rows(ws, 4, rows, fmt_map={4: currency_fmt()})

    # Gold/silver/bronze top 3
    medals = [("FFD700", CUSTOMER_DARK), ("C0C0C0", CUSTOMER_DARK), ("CD7F32", WHITE)]
    for i, (fill, font_col) in enumerate(medals):
        for c in range(1, 5):
            ws.cell(4 + i, c).fill = cell_fill(fill)
            ws.cell(4 + i, c).font = Font(name="Arial", size=10, bold=True, color=font_col)

    set_col_widths(ws, {"A": 8, "B": 18, "C": 16, "D": 22})


def build_anomalies(wb, anomaly_df):
    ws = wb.create_sheet("Anomaly Report")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:G1")
    ws["A1"].value     = "Anomaly Report — High-Value Outlier Transactions (>3σ)"
    ws["A1"].font      = Font(name="Arial", size=13, bold=True, color=WHITE)
    ws["A1"].fill      = cell_fill(RED_ALERT)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 24

    ws.merge_cells("A2:G2")
    ws["A2"].value     = "Transactions exceeding 3 standard deviations from segment mean. Requires compliance review."
    ws["A2"].font      = Font(name="Arial", size=10, italic=True, color="AA0000")
    ws["A2"].alignment = Alignment(horizontal="left")
    ws.row_dimensions[2].height = 18

    cols = ["Txn ID", "Date", "Customer", "Segment", "Branch", "Type", "Amount (RM)"]
    write_table_header(ws, 4, cols, fill_color=RED_ALERT)

    rows = []
    for _, r in anomaly_df.iterrows():
        rows.append([
            r["transaction_id"], r["date"], r["customer_id"],
            r["segment"], r["branch"], r["transaction_type"], r["amount"]
        ])
    write_data_rows(ws, 5, rows, fmt_map={7: currency_fmt()})

    # Red-tint amount column
    for row in range(5, 5 + len(rows)):
        ws.cell(row, 7).fill = cell_fill("FFE0E0")
        ws.cell(row, 7).font = Font(name="Arial", size=10, bold=True, color=RED_ALERT)

    set_col_widths(ws, {"A": 14, "B": 14, "C": 14, "D": 20, "E": 18, "F": 22, "G": 18})


# ── MAIN REPORT FUNCTION ──────────────────────────────────────────────────────

def generate_report(results: dict, output_path: str):
    wb = openpyxl.Workbook()

    build_cover(wb, results["kpi"])
    build_monthly(wb, results["monthly"])
    build_branch(wb, results["branch"])
    build_txn_breakdown(wb, results["by_type"], results["by_product"])
    build_segment(wb, results["by_segment"])
    build_top_customers(wb, results["top_customers"])
    build_anomalies(wb, results["anomalies"])

    # Tab colors
    tab_colors = {
        "Cover":                  CUSTOMER_YELLOW,
        "Monthly Trend":          "2E75B6",
        "Branch Analysis":        "375623",
        "Transaction Breakdown":  "7030A0",
        "Segment Analysis":       "2E75B6",
        "Top Customers":          "C09000",
        "Anomaly Report":         "C00000",
    }
    for sheet_name, color in tab_colors.items():
        if sheet_name in wb.sheetnames:
            wb[sheet_name].sheet_properties.tabColor = color

    wb.save(output_path)
    print(f"Report saved to: {output_path}")

